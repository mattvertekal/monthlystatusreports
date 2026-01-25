#!/usr/bin/env python3
"""
MSR Update Orchestrator
Coordinates updating all three MSRs (TO1, TO8, TO6) with timesheet data.

Supports two modes:
1. Auto mode: python update_msrs.py "Jan-26" (finds latest MSRs automatically)
2. Manual mode: python update_msrs.py --files to1.xlsx to8.xlsx to6.xlsx "Jan-26"
"""

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from agents.to1_updater import update_to1_msr
from agents.to8_updater import update_to8_msr
from agents.to6_updater import update_to6_msr
from utils.timesheet_parser import parse_timesheet_csv, get_timesheets_for_month
from utils.date_finder import parse_month_input, find_month_column, format_month_display
from openpyxl import load_workbook

# Default MSR directory (in project folder)
MSR_BASE_DIR = Path(__file__).parent.parent / "MSRs"
TEMPLATES_DIR = MSR_BASE_DIR / "templates"
COMPLETED_DIR = MSR_BASE_DIR / "completed"

# MSR file patterns (for finding files)
MSR_PATTERNS = {
    "TO1": ["TO1", "Athena TO1"],
    "TO8": ["TO8", "Athena TO8"],
    "TO6": ["TO6", "Athena TO6"]
}

# MSR output naming conventions
# TO1: "Athena TO1 Vertekal MSR Dec 2025.xlsx"
# TO8: "Athena TO8-Vertekal MSR_2026.01.xlsx"
# TO6: "Athena TO6 Vertekal MSR Opt3 December 2025.xlsx"
def get_msr_output_filename(msr_type: str, year: int, month: int) -> str:
    """Generate output filename matching the established naming convention."""
    from datetime import datetime
    dt = datetime(year, month, 1)

    if msr_type == "TO1":
        # Format: "Athena TO1 Vertekal MSR Dec 2025.xlsx"
        return f"Athena TO1 Vertekal MSR {dt.strftime('%b')} {year}.xlsx"
    elif msr_type == "TO8":
        # Format: "Athena TO8-Vertekal MSR_2026.01.xlsx"
        return f"Athena TO8-Vertekal MSR_{year}.{month:02d}.xlsx"
    elif msr_type == "TO6":
        # Format: "Athena TO6 Vertekal MSR Opt3 December 2025.xlsx"
        return f"Athena TO6 Vertekal MSR Opt3 {dt.strftime('%B')} {year}.xlsx"
    else:
        return f"{msr_type}_MSR_{year}-{month:02d}.xlsx"


def get_previous_month(year: int, month: int) -> Tuple[int, int]:
    """Get the previous month's year and month."""
    if month == 1:
        return year - 1, 12
    return year, month - 1


def get_month_folder_name(year: int, month: int) -> str:
    """Get folder name like '01-Jan' for a given month."""
    from datetime import datetime
    dt = datetime(year, month, 1)
    return dt.strftime("%m-%b")


def find_latest_msr(msr_type: str, before_year: int, before_month: int) -> Optional[Path]:
    """
    Find the most recent MSR file for a given type.

    Searches completed folders in reverse chronological order,
    then falls back to templates.

    Args:
        msr_type: "TO1", "TO8", or "TO6"
        before_year: Target year (search for months before this)
        before_month: Target month (search for months before this)

    Returns:
        Path to the most recent MSR file, or None if not found
    """
    patterns = MSR_PATTERNS.get(msr_type, [msr_type])

    # Search completed folders in reverse order
    year = before_year
    month = before_month - 1
    if month < 1:
        month = 12
        year -= 1

    # Search up to 24 months back
    for _ in range(24):
        folder_name = get_month_folder_name(year, month)
        folder_path = COMPLETED_DIR / str(year) / folder_name

        if folder_path.exists():
            for file in folder_path.iterdir():
                if file.suffix.lower() in ['.xlsx', '.xlsm', '.xlsb']:
                    file_upper = file.name.upper()
                    if any(p.upper() in file_upper for p in patterns):
                        return file

        # Go to previous month
        month -= 1
        if month < 1:
            month = 12
            year -= 1

    # Fall back to templates
    if TEMPLATES_DIR.exists():
        for file in TEMPLATES_DIR.iterdir():
            if file.suffix.lower() in ['.xlsx', '.xlsm', '.xlsb']:
                file_upper = file.name.upper()
                if any(p.upper() in file_upper for p in patterns):
                    return file

    return None


def find_all_msrs(target_year: int, target_month: int) -> Dict[str, Optional[Path]]:
    """Find the latest MSR file for each type."""
    return {
        "TO1": find_latest_msr("TO1", target_year, target_month),
        "TO8": find_latest_msr("TO8", target_year, target_month),
        "TO6": find_latest_msr("TO6", target_year, target_month)
    }


def update_all_msrs(
    timesheet_data: Dict[str, Dict[str, float]],
    to1_msr: str,
    to8_msr: str,
    to6_msr: str,
    month_str: str,
    output_dir: str = None
) -> Dict[str, any]:
    """
    Update all MSRs with timesheet data for the specified month.

    Args:
        timesheet_data: Parsed timesheet data (from CSV or API)
        to1_msr: Path to TO1 MSR file
        to8_msr: Path to TO8 MSR file
        to6_msr: Path to TO6 MSR file
        month_str: Month to update (e.g., "Jan-26", "February 2026")
        output_dir: Directory to save updated MSRs (default: completed/YYYY/MM-Mon/)

    Returns:
        Dictionary with results from all updates
    """
    print("="*80)
    print("MSR Update Process Starting")
    print("="*80)

    # Parse month
    try:
        target_year, target_month = parse_month_input(month_str)
        month_display = format_month_display(target_year, target_month)
        print(f"\n Target Month: {month_display} ({target_year}-{target_month:02d})")
    except ValueError as e:
        print(f"\nERROR: {e}")
        return None

    # Report timesheet data
    total_employees = len(timesheet_data)
    total_hours = sum(sum(codes.values()) for codes in timesheet_data.values())
    print(f" Found {total_employees} employees with {total_hours:.2f} billable hours")

    # Set output directory
    if output_dir is None:
        folder_name = get_month_folder_name(target_year, target_month)
        output_dir = COMPLETED_DIR / str(target_year) / folder_name
    else:
        output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f" Output directory: {output_dir}")

    # Load config for settings
    config_dir = Path(__file__).parent / 'config'
    with open(config_dir / 'msr_settings.json') as f:
        settings = json.load(f)

    results = {}

    # Update TO1
    print("\n" + "-"*80)
    print("Updating TO1 MSR...")
    print(f"   Source: {to1_msr}")
    print("-"*80)
    try:
        wb1 = load_workbook(to1_msr, data_only=True)
        ws1 = wb1[settings['TO1']['sheet_name']]
        date_row = settings['TO1']['date_header_row']
        col1 = find_month_column(ws1, date_row, target_year, target_month)
        wb1.close()

        if col1:
            print(f"   Found column: {col1}")
            output1 = output_dir / get_msr_output_filename("TO1", target_year, target_month)
            result1 = update_to1_msr(to1_msr, timesheet_data, col1, str(output1))
            results['TO1'] = result1
            print(f"   Updated {result1['total_hours']:.2f} hours")
            print(f"   Saved: {output1}")
        else:
            print(f"   Could not find column for {month_display}")
            results['TO1'] = {'error': 'Column not found'}
    except Exception as e:
        print(f"   Error: {e}")
        results['TO1'] = {'error': str(e)}

    # Update TO8
    print("\n" + "-"*80)
    print("Updating TO8 MSR...")
    print(f"   Source: {to8_msr}")
    print("-"*80)
    try:
        wb8 = load_workbook(to8_msr, data_only=True)
        ws8 = wb8['CLIN 0001AA']
        date_row = settings['TO8']['sheets']['CLIN 0001AA']['date_header_row']
        col8 = find_month_column(ws8, date_row, target_year, target_month)
        wb8.close()

        if col8:
            print(f"   Found column: {col8}")
            output8 = output_dir / get_msr_output_filename("TO8", target_year, target_month)
            result8 = update_to8_msr(to8_msr, timesheet_data, col8, str(output8))
            results['TO8'] = result8
            print(f"   CLIN 0001AA: {result8['clin_0001aa_hours']:.2f} hours")
            print(f"   CLIN 0002AA: {result8['clin_0002aa_hours']:.2f} hours")
            print(f"   Total: {result8['total_hours']:.2f} hours")
            print(f"   Saved: {output8}")
        else:
            print(f"   Could not find column for {month_display}")
            results['TO8'] = {'error': 'Column not found'}
    except Exception as e:
        print(f"   Error: {e}")
        results['TO8'] = {'error': str(e)}

    # Update TO6
    print("\n" + "-"*80)
    print("Updating TO6 MSR...")
    print(f"   Source: {to6_msr}")
    print("-"*80)
    try:
        wb6 = load_workbook(to6_msr, data_only=True)
        ws6 = wb6[settings['TO6']['sheet_name']]
        date_row = settings['TO6']['date_header_row']
        col6 = find_month_column(ws6, date_row, target_year, target_month)
        wb6.close()

        if col6:
            print(f"   Found column: {col6}")
            output6 = output_dir / get_msr_output_filename("TO6", target_year, target_month)
            result6 = update_to6_msr(to6_msr, timesheet_data, col6, str(output6))
            results['TO6'] = result6
            print(f"   Updated {result6['total_hours']:.2f} hours")
            print(f"   Saved: {output6}")
        else:
            print(f"   Could not find column for {month_display}")
            results['TO6'] = {'error': 'Column not found'}
    except Exception as e:
        print(f"   Error: {e}")
        results['TO6'] = {'error': str(e)}

    # Summary
    print("\n" + "="*80)
    print("MSR Update Complete")
    print("="*80)

    successful = sum(1 for r in results.values() if 'error' not in r)
    print(f"\nSuccessfully updated: {successful}/3 MSRs")
    print(f"Files saved to: {output_dir}")

    return results


def main():
    parser = argparse.ArgumentParser(
        description='Update MSRs with timesheet data',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Auto-find latest MSRs and update (recommended):
  python update_msrs.py "Jan-26"

  # Specify MSR files manually:
  python update_msrs.py --files to1.xlsx to8.xlsx to6.xlsx "Jan-26"

  # Use CSV instead of TSheets API:
  python update_msrs.py --csv timesheet.csv "Jan-26"
        """
    )

    parser.add_argument('month', help='Month to update (e.g., "Jan-26", "February 2026")')
    parser.add_argument('--files', nargs=3, metavar=('TO1', 'TO8', 'TO6'),
                        help='Manually specify MSR files')
    parser.add_argument('--csv', type=str, default=None,
                        help='Use CSV file instead of TSheets API')
    parser.add_argument('--output', '-o', type=str, default=None,
                        help='Output directory for updated MSRs')

    args = parser.parse_args()

    # Parse month first
    try:
        target_year, target_month = parse_month_input(args.month)
    except ValueError as e:
        print(f"Error parsing month: {e}")
        sys.exit(1)

    # Find or use specified MSR files
    if args.files:
        to1_msr, to8_msr, to6_msr = args.files
    else:
        print("Searching for latest MSR files...")
        msr_files = find_all_msrs(target_year, target_month)

        missing = [k for k, v in msr_files.items() if v is None]
        if missing:
            print(f"\nError: Could not find MSR files for: {', '.join(missing)}")
            print(f"\nPlease either:")
            print(f"  1. Drop MSR files into {COMPLETED_DIR}/YYYY/MM-Mon/")
            print(f"  2. Drop template files into {TEMPLATES_DIR}/")
            print(f"  3. Use --files to specify paths manually")
            sys.exit(1)

        to1_msr = str(msr_files['TO1'])
        to8_msr = str(msr_files['TO8'])
        to6_msr = str(msr_files['TO6'])

        print(f"  TO1: {to1_msr}")
        print(f"  TO8: {to8_msr}")
        print(f"  TO6: {to6_msr}")

    # Get timesheet data
    if args.csv:
        print(f"\nParsing timesheet CSV: {args.csv}")
        timesheet_data = parse_timesheet_csv(args.csv)
    else:
        print(f"\nFetching timesheets from TSheets API for {args.month}...")
        timesheet_data = get_timesheets_for_month(target_year, target_month)

    # Run the update
    update_all_msrs(
        timesheet_data=timesheet_data,
        to1_msr=to1_msr,
        to8_msr=to8_msr,
        to6_msr=to6_msr,
        month_str=args.month,
        output_dir=args.output
    )


if __name__ == "__main__":
    main()
