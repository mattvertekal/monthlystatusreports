"""
TO1 MSR Updater Agent
Updates Athena TO1 Extension Period MSR with timesheet data.
"""

import json
import copy
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict

import sys
sys.path.append(str(Path(__file__).parent.parent))

from utils.date_finder import find_month_column
from utils.timesheet_parser import parse_timesheet_csv


def update_to1_msr(
    msr_path: str,
    timesheet_data: Dict[str, Dict[str, float]],
    target_column: int,
    output_path: str
) -> Dict[str, any]:
    """
    Update TO1 MSR with timesheet hours.

    Args:
        msr_path: Path to the TO1 MSR file
        timesheet_data: Parsed timesheet data from timesheet_parser
        target_column: Column number to update
        output_path: Path to save updated MSR

    Returns:
        Dictionary with update summary
    """
    # Load configs
    config_dir = Path(__file__).parent.parent / 'config'
    with open(config_dir / 'employee_mappings.json') as f:
        mappings = json.load(f)
    with open(config_dir / 'msr_settings.json') as f:
        settings = json.load(f)

    # Load workbook
    wb = load_workbook(msr_path)
    ws = wb['Extension Period MSR']

    to1_settings = settings['TO1']
    status_row = to1_settings['status_row']
    total_row = to1_settings['total_row']

    # Get fill color from previous column (column - 1)
    prev_col = target_column - 1

    # Update status row
    status_fill = copy.copy(ws.cell(row=status_row, column=prev_col).fill)
    ws.cell(row=status_row, column=target_column).value = "Actual"
    ws.cell(row=status_row, column=target_column).fill = status_fill

    # Track updates
    updates = []
    total_hours = 0.0

    # Process each employee
    for emp_name, employee_info in mappings['employees'].items():
        if 'TO1' not in employee_info['msrs']:
            continue

        # Get timesheet hours for this employee
        emp_hours = timesheet_data.get(emp_name, {})

        # Update each charge code
        for charge_code, mapping in employee_info['charge_codes'].items():
            if mapping['msr'] != 'TO1':
                continue

            row = mapping['row']
            hours = emp_hours.get(charge_code, 0.0)

            # Get row fill from previous column
            row_fill = copy.copy(ws.cell(row=row, column=prev_col).fill)

            # Update cell
            ws.cell(row=row, column=target_column).value = hours
            ws.cell(row=row, column=target_column).fill = row_fill

            total_hours += hours
            updates.append({
                'employee': emp_name,
                'charge_code': charge_code,
                'row': row,
                'hours': hours
            })

    # Update total row
    total_fill = copy.copy(ws.cell(row=total_row, column=prev_col).fill)
    ws.cell(row=total_row, column=target_column).value = total_hours
    ws.cell(row=total_row, column=target_column).fill = total_fill

    # Save workbook
    wb.save(output_path)

    return {
        'msr': 'TO1',
        'total_hours': total_hours,
        'updates': updates,
        'output_path': output_path
    }


if __name__ == "__main__":
    # Test the updater
    print("TO1 Updater Agent - Ready")
