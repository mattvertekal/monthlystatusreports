"""
TO4 MSR Updater Agent
Updates Athena TO4 PIVOT MSR with timesheet data (both CLINs).
"""

import json
import copy
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict

import sys
sys.path.append(str(Path(__file__).parent.parent))

from utils.date_finder import find_month_column
from utils.timesheet_parser import parse_timesheet_csv


def update_to4_msr(
    msr_path: str,
    timesheet_data: Dict[str, Dict[str, float]],
    target_column: int,
    output_path: str
) -> Dict[str, any]:
    """
    Update TO4 MSR with timesheet hours for both CLINs.

    Args:
        msr_path: Path to the TO4 MSR file
        timesheet_data: Parsed timesheet data from timesheet_parser
        target_column: Column number to update
        output_path: Path to save updated MSR

    Returns:
        Dictionary with update summary
    """
    # Load configs
    config_dir = Path(__file__).parent.parent / 'config'
    with open(config_dir / 'employee_mappings.json') as f:
        mappings = json.load(f)
    with open(config_dir / 'msr_settings.json') as f:
        settings = json.load(f)

    # Load workbook
    wb = load_workbook(msr_path)

    to4_settings = settings['TO4']
    prev_col = target_column - 1

    # Track updates
    all_updates = {
        'CLIN 0001AD': [],
        'CLIN 0002AD': []
    }

    # Process each CLIN sheet
    for sheet_name, sheet_settings in to4_settings['sheets'].items():
        ws = wb[sheet_name]

        status_row = sheet_settings['status_row']
        fill_range = sheet_settings['fill_range']

        # Get status fill from previous column
        status_fill = copy.copy(ws.cell(row=status_row, column=prev_col).fill)

        # Update status and fill entire column range
        ws.cell(row=status_row, column=target_column).value = "Actual"
        for row in range(fill_range['start'], fill_range['end'] + 1):
            ws.cell(row=row, column=target_column).fill = status_fill

        # Process each employee for this sheet
        for emp_name, employee_info in mappings['employees'].items():
            if 'TO4' not in employee_info['msrs']:
                continue

            # Get timesheet hours for this employee
            emp_hours = timesheet_data.get(emp_name, {})

            # Update each charge code that belongs to this sheet
            for charge_code, mapping in employee_info['charge_codes'].items():
                if mapping['sheet'] != sheet_name:
                    continue

                row = mapping['row']
                hours = emp_hours.get(charge_code, 0.0)

                # Update cell value (fill was already set above)
                ws.cell(row=row, column=target_column).value = hours

                all_updates[sheet_name].append({
                    'employee': emp_name,
                    'charge_code': charge_code,
                    'row': row,
                    'hours': hours
                })

    # Save workbook
    wb.save(output_path)

    # Calculate totals
    total_clin1 = sum(u['hours'] for u in all_updates['CLIN 0001AD'])
    total_clin2 = sum(u['hours'] for u in all_updates['CLIN 0002AD'])

    return {
        'msr': 'TO4',
        'total_hours': total_clin1 + total_clin2,
        'clin_0001ad_hours': total_clin1,
        'clin_0002ad_hours': total_clin2,
        'updates': all_updates,
        'output_path': output_path
    }


if __name__ == "__main__":
    # Test the updater
    print("TO4 Updater Agent - Ready")
