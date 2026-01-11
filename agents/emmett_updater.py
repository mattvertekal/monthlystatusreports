"""
Emmett/Magni HA Updater Agent
Updates Vertekal subcontract WSR with timesheet data from TSheets.

This handles the Emmett (Magni HA) project for the Vertekal -> Booz Allen subcontract.
"""

import json
import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Dict, Optional

import sys
sys.path.append(str(Path(__file__).parent.parent))

from utils.date_finder import find_month_column
from utils.timesheet_parser import get_timesheets_for_month


def update_emmett_wsr(
    wsr_path: str,
    timesheet_data: Dict[str, Dict[str, float]],
    target_column: int,
    output_path: str
) -> Dict[str, any]:
    """
    Update Emmett WSR with timesheet hours.

    Args:
        wsr_path: Path to the WSR file (.xlsb or .xlsx)
        timesheet_data: Parsed timesheet data from TSheets
        target_column: Column number to update
        output_path: Path to save updated WSR

    Returns:
        Dictionary with update summary
    """
    # Load configs
    config_dir = Path(__file__).parent.parent / 'config'
    with open(config_dir / 'employee_mappings.json') as f:
        mappings = json.load(f)
    with open(config_dir / 'msr_settings.json') as f:
        settings = json.load(f)

    # Load workbook
    wb = load_workbook(wsr_path)
    ws = wb[settings['Emmett']['sheet_name']]

    emmett_settings = settings['Emmett']
    status_row = emmett_settings['status_row']
    highlight_color = emmett_settings.get('highlight_color', [217, 226, 243])

    # Get fill color from previous column
    prev_col = target_column - 1

    # Update status row (Estimate -> Actual)
    status_fill = copy.copy(ws.cell(row=status_row, column=prev_col).fill)
    ws.cell(row=status_row, column=target_column).value = "Actual"
    ws.cell(row=status_row, column=target_column).fill = status_fill

    # Create highlight fill
    highlight_fill = PatternFill(
        start_color=f'{highlight_color[0]:02X}{highlight_color[1]:02X}{highlight_color[2]:02X}',
        end_color=f'{highlight_color[0]:02X}{highlight_color[1]:02X}{highlight_color[2]:02X}',
        fill_type='solid'
    )

    # Track updates
    updates = []
    total_hours = 0.0

    # Process each employee
    for emp_name, employee_info in mappings['employees'].items():
        if 'Emmett' not in employee_info.get('msrs', []):
            continue

        # Get timesheet hours for this employee
        emp_hours = timesheet_data.get(emp_name, {})

        # Update each charge code
        for charge_code, mapping in employee_info['charge_codes'].items():
            if mapping['msr'] != 'Emmett':
                continue

            row = mapping['row']
            hours = emp_hours.get(charge_code, 0.0)

            # Copy format from previous column
            row_fill = copy.copy(ws.cell(row=row, column=prev_col).fill)

            # Update cell with hours
            cell = ws.cell(row=row, column=target_column)
            cell.value = hours

            # Apply highlight color
            cell.fill = highlight_fill

            total_hours += hours
            updates.append({
                'employee': emp_name,
                'charge_code': charge_code,
                'row': row,
                'hours': hours
            })

    # Save workbook
    wb.save(output_path)
    wb.close()

    return {
        'msr': 'Emmett',
        'total_hours': total_hours,
        'updates': updates,
        'output_path': output_path
    }


def update_emmett_from_api(
    wsr_path: str,
    year: int,
    month: int,
    output_path: Optional[str] = None
) -> Dict[str, any]:
    """
    Update Emmett WSR by fetching hours directly from TSheets API.

    Args:
        wsr_path: Path to the WSR file
        year: Year to update
        month: Month to update (1-12)
        output_path: Path to save (default: same dir with _UPDATED suffix)

    Returns:
        Update result dictionary
    """
    # Load settings
    config_dir = Path(__file__).parent.parent / 'config'
    with open(config_dir / 'msr_settings.json') as f:
        settings = json.load(f)

    # Fetch timesheet data
    print(f"Fetching TSheets data for {year}-{month:02d}...")
    timesheet_data = get_timesheets_for_month(year, month)

    print(f"Found hours for {len(timesheet_data)} employees:")
    for emp, codes in timesheet_data.items():
        total = sum(codes.values())
        print(f"  {emp}: {total:.2f} hrs")

    # Find target column
    wb = load_workbook(wsr_path, data_only=True)
    ws = wb[settings['Emmett']['sheet_name']]
    date_row = settings['Emmett']['date_header_row']

    target_col = find_month_column(ws, date_row, year, month)
    wb.close()

    if not target_col:
        return {
            'error': f'Could not find column for {year}-{month:02d}',
            'timesheet_data': timesheet_data
        }

    print(f"Target column: {target_col}")

    # Set output path
    if output_path is None:
        wsr_file = Path(wsr_path)
        output_path = str(wsr_file.parent / f"{wsr_file.stem}_{year}-{month:02d}_UPDATED{wsr_file.suffix}")

    # Update WSR
    result = update_emmett_wsr(wsr_path, timesheet_data, target_col, output_path)

    print(f"\nUpdated {result['total_hours']:.2f} total hours")
    print(f"Saved to: {output_path}")

    return result


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("Usage: python emmett_updater.py <wsr_file> <month>")
        print("Example: python emmett_updater.py ../wsr/Vertekal_WSR.xlsx Jan-26")
        sys.exit(1)

    wsr_path = sys.argv[1]
    month_str = sys.argv[2]

    # Parse month
    from utils.date_finder import parse_month_input
    year, month = parse_month_input(month_str)

    # Run update
    result = update_emmett_from_api(wsr_path, year, month)

    if 'error' in result:
        print(f"Error: {result['error']}")
        sys.exit(1)

    print("\nUpdate complete!")
    for update in result['updates']:
        print(f"  {update['employee']}: {update['hours']:.2f} hrs (row {update['row']})")
