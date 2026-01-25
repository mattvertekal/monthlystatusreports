"""
TO8 MSR Updater Agent
Updates Athena TO8 MSR with timesheet data (both CLINs).
"""

import json
import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Dict

import sys
sys.path.append(str(Path(__file__).parent.parent))

from utils.date_finder import find_month_column

# Light blue fill for "Actual" columns (matches TO4 style)
ACTUAL_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")


def update_to8_msr(
    msr_path: str,
    timesheet_data: Dict[str, Dict[str, float]],
    target_column: int,
    output_path: str
) -> Dict[str, any]:
    """
    Update TO8 MSR with timesheet hours for both CLINs.

    Args:
        msr_path: Path to the TO8 MSR file
        timesheet_data: Parsed timesheet data from timesheet_parser
        target_column: Column number to update
        output_path: Path to save updated MSR

    Returns:
        Dictionary with update summary
    """
    # Load configs
    config_dir = Path(__file__).parent.parent / 'config'
    with open(config_dir / 'employee_mappings.json') as f:
        mappings = json.load(f)
    with open(config_dir / 'msr_settings.json') as f:
        settings = json.load(f)

    # Load workbook
    wb = load_workbook(msr_path)

    to8_settings = settings['TO8']
    prev_col = target_column - 1

    # Track updates
    all_updates = {
        'CLIN 0001AA': [],
        'CLIN 0002AA': []
    }

    # Process each CLIN sheet
    for sheet_name, sheet_settings in to8_settings['sheets'].items():
        ws = wb[sheet_name]

        status_row = sheet_settings['status_row']
        fill_range = sheet_settings['fill_range']

        # Try to get fill from previous column, fall back to light blue
        prev_fill = ws.cell(row=status_row, column=prev_col).fill
        if prev_fill.fill_type and prev_fill.fill_type != 'none':
            status_fill = copy.copy(prev_fill)
        else:
            status_fill = ACTUAL_FILL

        # Update status to "Actual" and apply fill to the range
        ws.cell(row=status_row, column=target_column).value = "Actual"
        for row in range(fill_range['start'], fill_range['end'] + 1):
            ws.cell(row=row, column=target_column).fill = status_fill

        # Process each employee for this sheet
        for emp_name, employee_info in mappings['employees'].items():
            if 'TO8' not in employee_info.get('msrs', []):
                continue

            # Get timesheet hours for this employee
            emp_hours = timesheet_data.get(emp_name, {})

            # Update each charge code that belongs to this sheet
            for charge_code, mapping in employee_info['charge_codes'].items():
                if mapping.get('msr') != 'TO8':
                    continue
                if mapping['sheet'] != sheet_name:
                    continue

                row = mapping['row']
                hours = emp_hours.get(charge_code, 0.0)

                # Update cell value (fill was already set above)
                ws.cell(row=row, column=target_column).value = hours

                all_updates[sheet_name].append({
                    'employee': emp_name,
                    'charge_code': charge_code,
                    'row': row,
                    'hours': hours
                })

    # Save workbook
    wb.save(output_path)

    # Calculate totals
    total_clin1 = sum(u['hours'] for u in all_updates['CLIN 0001AA'])
    total_clin2 = sum(u['hours'] for u in all_updates['CLIN 0002AA'])

    return {
        'msr': 'TO8',
        'total_hours': total_clin1 + total_clin2,
        'clin_0001aa_hours': total_clin1,
        'clin_0002aa_hours': total_clin2,
        'updates': all_updates,
        'output_path': output_path
    }


if __name__ == "__main__":
    # Test the updater
    print("TO8 Updater Agent - Ready")
